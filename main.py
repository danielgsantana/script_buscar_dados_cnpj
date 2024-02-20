from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import funcoes
import re
import time


while True:
    try:
        print("Automação LDR")
        print("---------------------------------------------------")
        print("Escolha uma das opções abaixo:")
        print("1 - Para caso sua planilha conter os nomes da empresa")
        print("2 - Para caso sua planilha seja de CNPJ")
        entrada = int(input("-> "))

        if entrada == 1 or entrada == 2:
            break 
        else:
            print("---------------------------------------------------")
            print("Por favor, insira apenas o número 1 ou 2.")
            print("---------------------------------------------------")
            time.sleep(1)
    except ValueError:
        print("---------------------------------------------------")
        print("Entrada inválida. Por favor, insira apenas números.")
        print("---------------------------------------------------")
        time.sleep(1)


if entrada == 1:

    #importando a planilha (INSERIR O NOME DO ARQUIVO EXCEL QUE CONTENHA OS NOMES DAS EMPRESAS)
    planilha = load_workbook("planilha inicial.xlsx")
    
    aba_ativa = planilha['principal']

    # Percorrendo a planilha para adicionar na lista
    empresa_nome_list = funcoes.capturar_cnpj(aba_ativa)

    # Inicia o uma instância do google webdriver
    service = Service()
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)

    # Acessando a URL
    url = 'https://cnpj.linkana.com/'

    driver.get(url)

    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="q"]').click()
    driver.find_element(By.XPATH, '//*[@id="q"]').send_keys('azship')
    driver.find_element(By.XPATH, '/html/body/div[1]/main/div[1]/div/div[2]/form/div/input[2]').click()

    aba_ativa = planilha.create_sheet(title='resultado gerado')

    aba_ativa['A1'] = 'NOME EMPRESA'
    aba_ativa['B1'] = 'CNPJ'
    aba_ativa['C1'] = 'RAZÃO SOCIAL'
    aba_ativa['D1'] = 'NOME FANTASIA'
    aba_ativa['E1'] = 'ABERTURA'
    aba_ativa['F1'] = 'CAPITAL'
    aba_ativa['G1'] = 'EMAIL'
    aba_ativa['H1'] = 'TELEFONE'
    aba_ativa['I1'] = 'MUNICIPIO'
    aba_ativa['J1'] = 'UF'
    aba_ativa['K1'] = 'CEP'
    aba_ativa['L1'] = 'CNAE PRINCIPAL'

    row = 2

    for item in empresa_nome_list:

        time.sleep(3)
        driver.find_element(By.XPATH, '//*[@id="q"]').click()
        driver.find_element(By.XPATH, '//*[@id="q"]').send_keys(item)
        driver.find_element(By.XPATH, '/html/body/div/main/div/form/div[2]/input').click()

        try:
            c = driver.find_element(By.XPATH, '/html/body/div/main/div/div/a/div/div/p[2]').text
            cnpj = re.sub(r'[^0-9]', '', c)

        except NoSuchElementException:
            cnpj = "CNPJ não encontrado"

        objeto_cnpj = funcoes.buscar_cnpj_api(cnpj)


        funcoes.adicionar_dados_planilha(item, objeto_cnpj, planilha, row)

        row += 1


if entrada == 2:

    #importando a planilha (INSERIR O NOME DO ARQUIVO EXCEL QUE CONTENHA OS NOMES DAS EMPRESAS)
    planilha = load_workbook("planilha inicial.xlsx")
    
    aba_ativa = planilha['principal']

    # Percorrendo a planilha para adicionar na lista
    cnpj_list = funcoes.capturar_cnpj(aba_ativa)

    empresas_result = funcoes.buscar_cnpj_api(cnpj_list)

    funcoes.adicionar_dados_planilha(cnpj_list, empresas_result)
